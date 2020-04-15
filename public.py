# -*- coding:utf-8 -*-
# -------------------------------------------------------------------------------
#  Name:        对excel 操作的封装
#  Purpose:     支持 2007及以后的 xecel 版本
#
#  Author:      zou.yw
#
#  Created:     2018/4/27
#  Modifiy:     2020/4/09
#  Copyright:   (c) zou.yw 2020
#  Licence:     MIT
# -------------------------------------------------------------------------------

from openpyxl import Workbook, load_workbook
import operator
import os
import re
import time,datetime

class public_methods(object):

    def read_excel(self,data=None, read_type='row',pack="dict"):
        """
        参数：
            data=excel文件路径，read_type=row/column按行读还是按列读，
            pack=dict/set/list 返回结果是字典还是集合（如果是集合，所有的工作簿单元格都放到一个集合中）
        功能：遍历excel文件，返回数据列表或集合（默认按行）
        返回：包含列表的字典（每个工作簿名称作为一个字典key，此字典的值是一个 包含 子列表 的 列表a，就是整个excel表格；）
             如果按行读：列表a 的每个子列表 就是 一行；
             如果按列读：列表a 的每个子列表 就是 一列；
             egg: {"sheet1":[[v1,v1-1], [v2,v2-1], [...]], "sheet2":[[...]], ...}
        备注：按列读取有BUG，取出来的数据不全，而且没有了 key
        """

        try:
            # 先获取此文件的有效行列
            excel_max_row_and_colum = self.get_excel_max_rows_and_colums(data)
            # print(excel_max_row_and_colum)
            # 打开文件
            workbook = load_workbook(data, read_only=True)
            # 获取所有sheet，返回列表，格式：[u'sheet1', u'sheet2']
            # workbook_sheets_list = workbook.sheetnames
            wb_sheets = {}
            tmp_list = []
            data_list = list()
            for sheet in workbook:
                max_row = excel_max_row_and_colum[sheet.title][0]
                max_colum = excel_max_row_and_colum[sheet.title][1]
                # print(sheet.title, max_row, max_colum)
                sheet_all_rows = []
                row_index = 1  # 当前的行数
                # 遍历工作簿所有单元格
                for row in sheet.rows:
                    # 遍历每个单元格
                    colum_index = 1  # 当前列
                    for cel in row:
                        if colum_index > max_colum:
                            break
                        else:
                            colum_index += 1
                        if cel.value:
                            data_list.append(cel.value)
                        else:
                            data_list.append("")

                    sheet_all_rows.append(data_list)
                    # 每循环一次，要清空一次（用于记录下一行）
                    data_list = []
                    if row_index > max_row:
                        break
                    else:
                        row_index += 1
                # 做一次矩阵式转换，让行列交互
                tmp_list = list(map(list, zip(*sheet_all_rows)))
                # print(tmp_list)
                if pack == "dict":
                    # 生成字典，每个sheet一个字典节点
                    wb_sheets[sheet.title] = tmp_list
        except Exception as err:
            print(err)
            return None
        if pack == "dict":
            return wb_sheets
        elif pack == "set":
            return set(data_list)
        else:
            return data_list


    def get_excel_max_rows_and_colums(self,file=None):
        """
        计算excel文件最大的行和列，因为有部分excel文档被设置了单元格格式，实际内容为空
        :param file:
        :return {sheet_name:[max_row, max_colum]}
        """
        max_cnt = {}
        # 打开文件
        try:
            # 只读文件，用 read_only 模式，省内存，而且速度快很多
            workbook = load_workbook(file, read_only=True)
            for sheet_name in workbook:
                empty_row_cnt = 0  # 计算多少空行
                max_row = 0 # 最大行数
                max_colum = 0  # 最大列
                for row in sheet_name.rows:
                    # print(row)
                    empty_flag = 0 # 空值标记,如果遍历一行后还是0，表示此行全空
                    colum_cnt = 0  # 计算列
                    for cel in row:
                        if cel.value:
                            colum_cnt += 1
                            if colum_cnt > max_colum:
                                max_colum = colum_cnt
                            empty_flag = 1
                    if not empty_flag:
                        empty_row_cnt += 1
                    else:
                        empty_row_cnt = 0
                        max_row += 1
                    if empty_row_cnt > 20:
                        # 连续空20行
                        break
                max_cnt[sheet_name.title] = [max_row, max_colum]
        except Exception as bug:
            print("we have a bug", bug)
        return max_cnt

    def write_excel(self,data=None,file_name=None,sheet_name=None,rows=0,cl=0):
        """
        功能：把数据写入 excel
        入参：data=要写入的数据（支持 列表、元组、字符串）; file_name=要写入的文件; rows,cl 开始写的行、列
        返回值：写入成功 Ture; 写入失败 False
        """

        if not isinstance(data,list):
            # return "The data to write , not a list"
            return None
        if not file_name:
            return None
        if os.path.exists(file_name):
            wb = load_workbook(file_name)
        else:
            # 新建一个工作簿（内存中），用来存放输出结果
            wb = Workbook()
            # 新建一张表
            ws_new = wb.active
            try:
                ws_new.append(data)
            except  Exception as bug:
                ws_new.append(['we have a problme. i have a bug'])
                ws_new.append([bug])
            finally:
                try:
                    wb.save(file_name)
                except  Exception as err:
                    print('无法保存文件，或许在编辑中')
        return


    def get_week_day(self,y=1970,m=1,d=1):
        """日期转换为星期"""
        t = time.strptime("%s - %s - %s" % (y, m, d), "%Y - %m - %d")
        y, m, d = t[0:3]
        week_day_dict = {
            0:'星期一',
            1:'星期二',
            2:'星期三',
            3:'星期四',
            4:'星期五',
            5:'星期六',
            6:'星期天',
        }
        date = datetime.datetime(y, m, d)
        day = date.weekday()
        return week_day_dict[day]

    def get_balance_day(self,y=1970,m=1,d=1,y1=1970,m1=1,d1=1):
        """获取两个日期相差天数
        入参：2个日期的年、月、日
        返回：相差的天数"""
        t = datetime.datetime.strptime("%s - %s - %s" % (y, m, d), "%Y - %m - %d")
        t1 = datetime.datetime.strptime("%s - %s - %s" % (y1, m1, d1), "%Y - %m - %d")
        return (t1 - t).days
